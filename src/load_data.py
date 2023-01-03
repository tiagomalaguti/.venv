from openpyxl import load_workbook
import win32com.client
import os
import csv

class Listas:
    def __init__(self):
        self.cpp = []
        self.descricao = []
        self.item = []
        self.contrato = []
        self.valor = []

    def load_data(self):
        wb = load_workbook("info.xlsx")
        ws = wb["Planilha1"]
        for row in ws.iter_rows(min_row=1):
            self.cpp.append(row[0].value)
            self.descricao.append(row[1].value)
            self.item.append(row[2].value)
            self.contrato.append(row[3].value)
            self.valor.append(row[4].value)
        print("upload concluido com sucesso!")
#_______________________________________________________________________________________________________________________________


class atualiza():
    def __init__(self):
        self.cpp_cont = []
        self.descricao_cont = []
        self.item_cont = []
        self.contrato_cont = []
        self.valor_cont = []

    def baixar_dados(self):
        
        base = Listas()

        SapGuiAuto = win32com.client.GetObject("SAPGUI")
        application = SapGuiAuto.GetScriptingEngine
        connection = application.Children(0)
        session = connection.Children(0)
        
        session.findById("wnd[0]/tbar[0]/okcd").text = "/nme3n"
        session.findById("wnd[0]/tbar[0]/btn[0]").press()
      
      #---------------------------------------------------
        session.findById("wnd[0]/usr/ctxtEN_EBELN-LOW").text = "4600176606"
        session.findById("wnd[0]/usr/ctxtLISTU").text = "ALLES_ALV"
        session.findById("wnd[0]/tbar[1]/btn[8]").press()

        current_path = os.getcwd()
        save_path = current_path + '/'
        
        session.findById("wnd[0]/tbar[1]/btn[45]").press()


        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[1]/usr/ctxtDY_PATH").text = save_path
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "SAP.txt"
        session.findById("wnd[1]/tbar[0]/btn[11]").press()
      # ---------------------------------------------------
        f = open('SAP.txt', 'r')


        aux = ""
        conteudo = f.readlines()
        for linha in conteudo:
            if (linha[0] == '|'):
                aux += linha
        arquivo = open('SAP.txt', 'w')
        arquivo.writelines(aux)
        arquivo.close()

      # ---------------------------------------------------
        with open('SAP.txt', 'r') as f:
            
            reader = csv.reader(f.readlines()[1:], delimiter='|')

           

            for row in reader:
                self.cpp_cont.append(row[1].replace(" ", ""))
                self.descricao_cont.append(row[2])
                self.item_cont.append(row[3])
                self.contrato_cont.append(row[4])
                self.valor_cont.append(row[5])

        base.load_data()
      # ----------------------------------------------
        status = 0
        for linha_contrato in range(len(self.cpp_cont)):
            for linha_base in range(len(base.cpp)):
                #print("rodando")
                #print( "numero total:"+ str(len(self.cpp_cont)) + " linha_contrato:" + str(linha_contrato) +  " cpp_base:" +base.cpp[linha_base] +"cpp_cont:" + self.cpp_cont[linha_contrato])

                if base.cpp[linha_base] == self.cpp_cont[linha_contrato]:
                    base.item[linha_base] = self.item_cont[linha_contrato]
                    base.descricao[linha_base] = self.descricao_cont[linha_contrato]
                    base.contrato[linha_base] = self.contrato_cont[linha_contrato]
                    base.valor[linha_base] = self.valor_cont[linha_contrato]
                    status = 1
                    break
                if (int(len(base.cpp)-1) == linha_base) and (status == 0) :
                    base.cpp.append(self.cpp_cont[linha_contrato])
                    base.item.append(self.item_cont[linha_contrato])
                    base.descricao.append(self.descricao_cont[linha_contrato])
                    base.contrato.append(self.contrato_cont[linha_contrato])
                    base.valor.append(self.valor_cont[linha_contrato])
                    status = 0
                    


        workbook = load_workbook("info.xlsx")

        worksheet = workbook["Planilha1"]
        

        for linha_base in range(len(base.cpp)):
            worksheet.cell(row=linha_base+1, column=1).value = str(base.cpp[linha_base])
            worksheet.cell(row=linha_base+1, column=2).value = str(base.descricao[linha_base])
            worksheet.cell(row=linha_base+1, column=3).value = str(base.item[linha_base])
            worksheet.cell(row=linha_base+1, column=4).value = str(base.contrato[linha_base])
            worksheet.cell(row=linha_base+1, column=5).value = str(base.valor[linha_base])
            
        workbook.save("info.xlsx")
